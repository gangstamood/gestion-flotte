"""
Import WLG26 — golfettes + attributions depuis le planning Excel.

Format Excel : 2 lignes par golfette (data + zone). La ligne des zones est
la source de vérité — la ligne 1.0 est incohérente (cf. G8 qui n'a aucun
marqueur 1.0 mais 5 segments de zones).

Pour chaque jour entre IN et OUT :
  - cellule zone remplie → attribution avec cette zone (whitespace normalisé)
  - cellule vide ou "x"  → attribution "SPARE"

Stratégie : remplace TOUTES les attributions_golfettes + remplace les
catégories_golfettes par les 4 nouveaux types (Cargo L, Cargo S, 4 places,
6 places).

Usage : .venv/bin/python3 import_wlg_golfettes.py
"""
import os
from datetime import date, timedelta
import openpyxl
import toml
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials

XLSX_PATH = '/Users/alan/Downloads/WLG - Golfette 2026 (1).xlsx'
SECRETS_PATH = os.path.join(os.path.dirname(__file__), '.streamlit', 'secrets.toml')

SHEET_NAME = ' REPARTITION GOLFETTES 2026'

# Plage des colonnes-dates de l'Excel
COL_FIRST = 6   # col F = 2026-05-13
COL_LAST = 40   # col AN = 2026-06-16
DATE_FIRST = date(2026, 5, 13)

# Nouvelles catégories à appliquer (remplace l'ancienne table)
NEW_CATEGORIES = ['Cargo L', 'Cargo S', '4 places', '6 places']


def col_to_date(c):
    return DATE_FIRST + timedelta(days=c - COL_FIRST)


def connect():
    secrets = toml.load(SECRETS_PATH)
    creds = Credentials.from_service_account_info(
        dict(secrets['gcp_service_account']),
        scopes=['https://www.googleapis.com/auth/spreadsheets']
    )
    sid = secrets['google_sheets']['spreadsheet_id']
    svc = build('sheets', 'v4', credentials=creds)
    return svc, sid


def read_sheet(svc, sid, name):
    result = svc.spreadsheets().values().get(
        spreadsheetId=sid, range=f"{name}!A:Z"
    ).execute()
    rows = result.get('values', [])
    if not rows or len(rows) < 2:
        return [], rows[0] if rows else []
    headers = rows[0]
    data = [dict(zip(headers, row + [''] * max(0, len(headers) - len(row)))) for row in rows[1:]]
    return data, headers


def write_sheet(svc, sid, name, data, headers=None):
    svc.spreadsheets().values().clear(spreadsheetId=sid, range=f"{name}!A:Z").execute()
    if not data and not headers:
        return
    if headers is None:
        headers = list(data[0].keys())
    values = [headers] + [[str(row.get(h, '') or '') for h in headers] for row in data]
    svc.spreadsheets().values().update(
        spreadsheetId=sid, range=f"{name}!A1",
        valueInputOption='RAW', body={'values': values}
    ).execute()
    print(f"  ✅ {name} : {len(data)} lignes écrites")


def main():
    print("📖 Lecture du fichier Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Map merged-cell coordinates → top-left value (pour résoudre les fusions)
    merged_lookup = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_lookup[(r, c)] = top_val

    def cell_val(r, c):
        if (r, c) in merged_lookup:
            return merged_lookup[(r, c)]
        return ws.cell(row=r, column=c).value

    # Repère les golfettes (col A à partir de la ligne 15, une golfette toutes les 2 lignes)
    golfettes_rows = []
    for r in range(14, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().startswith('G'):
            golfettes_rows.append((r, str(v).strip()))

    print(f"  {len(golfettes_rows)} golfettes détectées (G1 → G{len(golfettes_rows)})")

    golfettes_to_upsert = []
    attributions_to_add = []
    zones_seen = set()

    for data_row, gid in golfettes_rows:
        zone_row = data_row + 1
        type_raw = ws.cell(row=data_row, column=2).value
        in_d = ws.cell(row=data_row, column=3).value
        out_d = ws.cell(row=data_row, column=4).value
        type_norm = str(type_raw).strip() if type_raw else ''
        if not hasattr(in_d, 'date') or not hasattr(out_d, 'date'):
            print(f"  ⚠️  {gid} : IN/OUT manquants, ignoré")
            continue
        in_d, out_d = in_d.date(), out_d.date()

        golfettes_to_upsert.append({
            'numero_serie': gid,
            'type': type_norm,
        })

        # Pour chaque colonne-date, créer attribution si jour ∈ [IN, OUT]
        for c in range(COL_FIRST, COL_LAST + 1):
            d = col_to_date(c)
            if d < in_d or d > out_d:
                continue
            raw = cell_val(zone_row, c)
            zone = ' '.join(str(raw).split()).strip() if raw else ''
            # "x" ou vide → SPARE
            if not zone or zone.lower() == 'x':
                zone = 'SPARE'
            zones_seen.add(zone)
            attributions_to_add.append({
                'numero_serie': gid,
                'service': zone,
                'date': d.strftime("%d/%m/%Y"),
                'date_fin': d.strftime("%d/%m/%Y"),
                'periode': 'Journée',
                'retourne': '',
            })

    print(f"  {len(golfettes_to_upsert)} golfettes à upsert")
    print(f"  {len(attributions_to_add)} attributions journalières")
    print(f"  {len(zones_seen)} zones distinctes")

    print("\n🔌 Connexion Google Sheets...")
    svc, sid = connect()

    # ── 1. Catégories : remplace par les 4 nouvelles ───────────────────────
    cats_data, cats_headers = read_sheet(svc, sid, 'categories_golfettes')
    headers = cats_headers if cats_headers else ['nom']
    new_cats = [{'nom': c} for c in NEW_CATEGORIES]
    write_sheet(svc, sid, 'categories_golfettes', new_cats, headers=headers)
    print(f"  ✅ categories_golfettes remplacées : {NEW_CATEGORIES}")

    # ── 2. Golfettes : upsert (préserve marque/autres champs existants) ────
    existing_golf, golf_headers = read_sheet(svc, sid, 'golfettes')
    if not golf_headers:
        golf_headers = ['numero_serie', 'type', 'marque']
    by_id = {g.get('numero_serie'): g for g in existing_golf}
    nb_new, nb_upd = 0, 0
    for g in golfettes_to_upsert:
        if g['numero_serie'] in by_id:
            by_id[g['numero_serie']]['type'] = g['type']
            nb_upd += 1
        else:
            entry = {h: '' for h in golf_headers}
            entry['numero_serie'] = g['numero_serie']
            entry['type'] = g['type']
            by_id[g['numero_serie']] = entry
            nb_new += 1
    write_sheet(svc, sid, 'golfettes', list(by_id.values()), headers=golf_headers)
    print(f"  ✅ {nb_new} golfettes ajoutées, {nb_upd} mises à jour")

    # ── 3. Attributions : wipe + reimport complet ─────────────────────────
    existing_attrs, attrs_headers = read_sheet(svc, sid, 'attributions_golfettes')
    if not attrs_headers:
        attrs_headers = ['numero_serie', 'service', 'date', 'date_fin', 'periode', 'retourne']
    write_sheet(svc, sid, 'attributions_golfettes', attributions_to_add, headers=attrs_headers)
    print(f"  ✅ {len(attributions_to_add)} attributions WLG"
          f" (remplacé {len(existing_attrs)} anciennes)")

    print("\n🎉 Import golfettes WLG26 terminé !")
    print(f"\n{'ID':5s} {'Type':12s} {'Période':25s} {'Nb jours'}")
    print("-" * 60)
    for g in golfettes_to_upsert:
        attrs = [a for a in attributions_to_add if a['numero_serie'] == g['numero_serie']]
        if attrs:
            print(f"{g['numero_serie']:5s} {g['type']:12s} "
                  f"{attrs[0]['date']} → {attrs[-1]['date']:12s} {len(attrs)}")

    print(f"\nZones distinctes ({len(zones_seen)}) :")
    for z in sorted(zones_seen):
        print(f"  - {z}")


if __name__ == '__main__':
    main()
